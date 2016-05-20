import openpyxl



def main():
	print('Creating new files....')

	getData("Accelerometer")
	getData("Gyroscope")

	#does not include sensors called pressure sensors
	getData("Barometer")

	print('Done')

def getData(sensor):
	 
	#wb with data
	data = openpyxl.load_workbook('data.xlsx')
	sheet = data.get_sheet_by_name('Sheet1')

	#wb with reference titles
	titles = openpyxl.load_workbook('reference.xlsx')
	title_sheet = titles.get_sheet_by_name('Sheet1')

	#wb to save data to
	results = openpyxl.Workbook()
	new_sheet = results.active
	sheet.title = sensor + ' Results'
	counter = 1;

	for row in range(1, sheet.max_row):
		new_sheet.cell(row = counter, column = row).value = title_sheet.cell(row = counter, column = row).value
	counter += 1;
	#print('Working...')

	#iterate through .xlsx file
	for row in range(1,sheet.max_row):
		data = str(sheet['C' + str(row)].value).lower()
		if sensor.lower() in data:
			for cell in range(1, row):
				new_sheet.cell(row = counter, column = cell).value = sheet.cell(row = row, column = cell).value
			counter = counter + 1;
	#print('Creating new file...')
	results.save(sensor + ' Data.xlsx')

	#print('Done')


main()