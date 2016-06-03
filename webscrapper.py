#webscrapper to get info from LinkedIn

#import block
import urllib2 as URL
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver


def findComments(co, row):
	#By-pass LinkedIn's scraper blocking
	urlopener = URL.build_opener()
	urlopener.addheaders = [('User-agent', 'Mozilla/5.0')]

	link = "https://www.linkedin.com/company/" + co
	html = urlopener.open(link).read()

	#render JS/iFrame
	driver = webdriver.Firefox()
	profile_link = link
	driver.get(profile_link)
	html = driver.page_source
	
	wb = openpyxl.Workbook()
	ws = wb.active

	#web crawler
	soup = BeautifulSoup(html, "lxml")

	cell = 'A' + str(row)
	#for txt in soup.find_all("li", "feed-item"):
	#	print(txt.get_text())
	
	activity = soup.find_all("li", "feed-item")
	for item in list(activity):
		print(activity[1])

	wb.save(co + ".xlsx")

def main():
	row = 0
	findComments("state_farm", 1)
	#findComments("amazon", row+1)

main()

'''
		ws['A' + str(row)] = soup.find_all("div", "annotated-body").get_text()
		print(soup.find("div", "annotated-body").get_text())
		ws['B' + str(row)] = soup.find_all("span", "share-body").get_text()
		print(soup.find("span", "share-body").get_text())
		ws['C' + str(row)] = soup.find_all("li","feed-like").get_text()
		print(soup.find("li","feed-like").get_text())
		ws['D' + str(row)] = soup.find_all("li","feed-comment").get_text()
		print(soup.find("li","feed-comment").get_text())
		ws['E' + str(row)] = soup.find_all("li","feed-share").get_text()
		print(soup.find("li","feed-share").get_text())
		row = row + 1
'''